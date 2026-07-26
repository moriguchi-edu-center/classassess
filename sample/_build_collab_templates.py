# -*- coding: utf-8 -*-
"""共同編集テンプレートを、Forms出力と同じSheet1形式で再生成する。"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent
FORMS_SAMPLE = ROOT / "学級力アンケート_サンプル_第1回.xlsx"
OUT_CLASS = ROOT / "共同編集テンプレート_数式集約版_教室用.xlsx"
OUT_DEMO = ROOT / "共同編集テンプレート_個人シート45_Sheet1集約.xlsx"

CHOICE = ["そう思わない", "あまりそう思わない", "そう思う", "とてもそう思う"]
# 個人シート入力例（01〜05）※1〜4
DEMO_SCORES = {
    "01": [2, 2, 3, 3, 3, 2, 3, 2, 2, 3, 3, 3, 3, 3, 2],
    "02": [2, 3, 4, 3, 3, 3, 2, 2, 3, 3, 2, 3, 3, 3, 2],
    "03": [3, 2, 1, 2, 2, 3, 3, 2, 3, 2, 3, 2, 3, 3, 3],
    "04": [3, 3, 3, 4, 3, 3, 3, 3, 4, 3, 3, 3, 2, 3, 3],
    "05": [4, 3, 3, 3, 4, 3, 4, 3, 3, 4, 3, 4, 3, 3, 3],
}


def load_forms_headers() -> list[str]:
    wb = load_workbook(FORMS_SAMPLE, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if headers[:5] != ["ID", "開始時刻", "完了時刻", "メール", "名前"]:
        raise SystemExit(f"Formsサンプル先頭列が想定と違います: {headers[:5]}")
    if len(headers) < 20:
        raise SystemExit(f"設問列が不足: {len(headers)}")
    return [str(h) for h in headers[:20]]


def thin_border() -> Border:
    s = Side(style="thin", color="B0BEC5")
    return Border(left=s, right=s, top=s, bottom=s)


def style_header(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F6F78")
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = thin_border()


def score_to_text(score: int | None) -> str | None:
    if score is None:
        return None
    if not 1 <= int(score) <= 4:
        raise ValueError(score)
    return CHOICE[int(score) - 1]


def text_formula(sheet_name: str, row: int) -> str:
    # 個人シートB列の1〜4 → Formsと同じ選択肢テキスト
    ref = f"'{sheet_name}'!B{row}"
    return (
        f'=IF({ref}="","",'
        f'IFERROR(CHOOSE({ref},"そう思わない","あまりそう思わない","そう思う","とてもそう思う"),""))'
    )


def build_personal_sheet(wb: Workbook, sid: str, questions: list[str], scores: list[int] | None) -> None:
    ws = wb.create_sheet(sid)
    ws["A1"] = f"出席番号 {sid} の回答シート"
    ws["A1"].font = Font(bold=True, size=14, color="1F6F78")
    ws["A2"] = "B列に 1〜4 の数字を入れてください（4=とてもそう思う … 1=そう思わない）"
    ws["A2"].font = Font(color="5A7179", size=10)
    ws["A3"] = "設問"
    ws["B3"] = "回答(1-4)"
    style_header(ws["A3"])
    style_header(ws["B3"])

    for i, q in enumerate(questions):
        r = 4 + i
        ws.cell(r, 1, q)
        ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(r, 1).border = thin_border()
        cell = ws.cell(r, 2)
        if scores is not None:
            cell.value = scores[i]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
        cell.fill = PatternFill("solid", fgColor="FFF8E7")

    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 12
    ws.row_dimensions[1].height = 22
    for r in range(4, 19):
        ws.row_dimensions[r].height = 36

    dv = DataValidation(type="list", formula1='"1,2,3,4"', allow_blank=True)
    dv.error = "1〜4を選んでください"
    dv.errorTitle = "入力エラー"
    ws.add_data_validation(dv)
    dv.add("B4:B18")


def build_sheet1(
    wb: Workbook,
    headers: list[str],
    *,
    use_formulas: bool,
    demo_values: dict[str, list[int]] | None = None,
) -> None:
    ws = wb.create_sheet("Sheet1", 0)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        style_header(cell)
    ws.row_dimensions[1].height = 48

    questions_start_col = 6  # F列
    for i in range(1, 46):
        sid = f"{i:02d}"
        r = i + 1
        ws.cell(r, 1, sid)
        # Forms互換のメタ列（空でよい）
        for c in range(2, 6):
            ws.cell(r, c, None)
            ws.cell(r, c).border = thin_border()
        ws.cell(r, 1).border = thin_border()
        ws.cell(r, 1).alignment = Alignment(horizontal="center")

        for qi in range(15):
            c = questions_start_col + qi
            cell = ws.cell(r, c)
            cell.border = thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if use_formulas:
                cell.value = text_formula(sid, 4 + qi)
            elif demo_values and sid in demo_values:
                cell.value = score_to_text(demo_values[sid][qi])
            else:
                cell.value = None

    widths = [8, 14, 14, 12, 10] + [28] * 15
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}46"


def build_howto(wb: Workbook) -> None:
    ws = wb.create_sheet("使い方", 1)
    lines = [
        "共同編集ブック（Formsと同じSheet1形式）",
        "",
        "■ ポイント",
        "・ツールが読むのは Sheet1 だけです。",
        "・Sheet1 の列構成は Microsoft Forms のExcel出力と同じです。",
        "  （ID / 開始時刻 / 完了時刻 / メール / 名前 / ①〜⑮の設問）",
        "・そのため、Forms出力でもこのブックでも、ツール側の読み方は同じです。",
        "",
        "■ 子ども（入力）",
        "1. 自分の出席番号シート（01〜45）を開く",
        "2. B列に 1〜4 を入れる（プルダウンあり）",
        "",
        "■ 先生（集約・読込）",
        "1. Excel / Excel Online でブックを開き、保存する（数式結果を確定）",
        "2. Sheet1 に Formsと同じ形の表が並ぶ（回答はテキスト）",
        "3. その xlsx を学級力ツールの「Excelを読み込む」へ登録",
        "",
        "※ 開始時刻・完了時刻・メール・名前は空欄で問題ありません。",
    ]
    for i, line in enumerate(lines, 1):
        ws.cell(i, 1, line)
        if i == 1:
            ws.cell(i, 1).font = Font(bold=True, size=14, color="1F6F78")
    ws.column_dimensions["A"].width = 78


def make_workbook(*, classroom: bool) -> Workbook:
    headers = load_forms_headers()
    questions = headers[5:]
    wb = Workbook()
    # remove default
    default = wb.active
    wb.remove(default)

    if classroom:
        build_sheet1(wb, headers, use_formulas=True)
    else:
        # デモ用: 01〜05は計算済みテキストを直書き（ブラウザでも読める）
        # 06〜45は数式（Excelで保存後に値が入る）
        build_sheet1(wb, headers, use_formulas=False, demo_values=DEMO_SCORES)
        ws = wb["Sheet1"]
        for i in range(6, 46):
            sid = f"{i:02d}"
            r = i + 1
            for qi in range(15):
                ws.cell(r, 6 + qi).value = text_formula(sid, 4 + qi)

    build_howto(wb)

    for i in range(1, 46):
        sid = f"{i:02d}"
        scores = None if classroom else DEMO_SCORES.get(sid)
        # 教室用も見た目確認しやすいよう 01 だけ空のまま、デモ用は例値
        if classroom:
            scores = None
        build_personal_sheet(wb, sid, questions, scores)

    return wb


def main() -> None:
    class_wb = make_workbook(classroom=True)
    class_wb.save(OUT_CLASS)
    print("wrote", OUT_CLASS.name)

    demo_wb = make_workbook(classroom=False)
    demo_wb.save(OUT_DEMO)
    print("wrote", OUT_DEMO.name)

    # 検証: Sheet1見出しがFormsと一致
    h_forms = load_forms_headers()
    for path in (OUT_CLASS, OUT_DEMO):
        wb = load_workbook(path, data_only=False)
        h = [c.value for c in wb["Sheet1"][1]]
        assert h == h_forms, (path.name, h, h_forms)
        print("header OK:", path.name)


if __name__ == "__main__":
    main()
